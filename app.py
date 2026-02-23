import csv
import os
import re
import logging
from datetime import date, datetime, timedelta
from io import BytesIO, StringIO
from urllib.request import urlopen

from flask import (
    Flask, Response, abort, jsonify,
    redirect, render_template, request,
    url_for, send_file
)

from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import (
    Image as RLImage, Paragraph,
    SimpleDocTemplate, Spacer,
    Table, TableStyle
)

# 🔐 Conexão centralizada no db.py
from db import DBConnection, get_conn, DB_DIR
from migrations import run_migrations

logging.basicConfig(level=logging.INFO)

# 🔥 Executa migrations após importar conexão
run_migrations()

BASE_DIR = os.path.dirname(os.path.abspath(__file__))

ADMIN_TOKEN = os.getenv("ALMOCO_ADMIN_TOKEN", "ifc-sbs")

app = Flask(__name__)

# ------------------------------------------------
# IMPORTAÇÃO DOS BLUEPRINTS
# ------------------------------------------------
from routes_main import bp_main
from routes_admin import bp_admin

app.register_blueprint(bp_main)
app.register_blueprint(bp_admin)



# ------------------------------------------------
# HEALTHCHECK COMPATÍVEL COM POSTGRES
# ------------------------------------------------
@app.get("/health_db")
def health_db():
    try:
        with get_conn() as conn:
            alunos_row = conn.execute(
                "SELECT COUNT(*) as total FROM alunos"
            ).fetchone()

            respostas_row = conn.execute(
                "SELECT COUNT(*) as total FROM respostas"
            ).fetchone()

            quadro_row = conn.execute(
                "SELECT COUNT(*) as total FROM quadro_importado"
            ).fetchone()

            # Compatível com dict ou tuple
            alunos = alunos_row["total"] if isinstance(alunos_row, dict) else alunos_row[0]
            respostas = respostas_row["total"] if isinstance(respostas_row, dict) else respostas_row[0]
            quadro = quadro_row["total"] if isinstance(quadro_row, dict) else quadro_row[0]

        return jsonify({
            "status": "ok",
            "database": "postgres" if os.getenv("DATABASE_URL") else "sqlite",
            "alunos": alunos,
            "respostas": respostas,
            "quadro_importado": quadro
        })

    except Exception as e:
        logging.error(f"[ALMOCO] ERRO health_db: {e}")
        return jsonify({
            "status": "erro",
            "erro": str(e)
        }), 500


# ------------------------------------------------
# BACKUP XLSX
# ------------------------------------------------
def write_backup_xlsx() -> None:
    backup_dir = DB_DIR / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)

    backup_path = backup_dir / f"almoco_backup_{date.today().isoformat()}.xlsx"

    with get_conn() as conn:
        respostas_rows = conn.execute("""
            SELECT id, nome, matricula, turma, data_almoco, intencao, criado_em
            FROM respostas
            ORDER BY data_almoco, turma, nome
        """).fetchall()

        alunos_rows = conn.execute("""
            SELECT matricula, nome, turma, atualizado_em
            FROM alunos
            ORDER BY turma, nome
        """).fetchall()

        quadro_rows = conn.execute("""
            SELECT turma, data_almoco, sim, atualizado_em
            FROM quadro_importado
            ORDER BY data_almoco, turma
        """).fetchall()

    workbook = Workbook()

    ws_respostas = workbook.active
    ws_respostas.title = "respostas"
    ws_respostas.append(["id", "nome", "matricula", "turma", "data_almoco", "intencao", "criado_em"])
    for row in respostas_rows:
        ws_respostas.append([
            row["id"],
            row["nome"],
            row["matricula"],
            row["turma"],
            row["data_almoco"],
            row["intencao"],
            row["criado_em"]
        ])

    ws_alunos = workbook.create_sheet("alunos")
    ws_alunos.append(["matricula", "nome", "turma", "atualizado_em"])
    for row in alunos_rows:
        ws_alunos.append([
            row["matricula"],
            row["nome"],
            row["turma"],
            row["atualizado_em"]
        ])

    ws_quadro = workbook.create_sheet("quadro_importado")
    ws_quadro.append(["turma", "data_almoco", "sim", "atualizado_em"])
    for row in quadro_rows:
        ws_quadro.append([
            row["turma"],
            row["data_almoco"],
            row["sim"],
            row["atualizado_em"]
        ])

    workbook.save(backup_path)


# ------------------------------------------------
# START
# ------------------------------------------------
if __name__ == "__main__":
    app.run(
        host="0.0.0.0",
        port=int(os.getenv("PORT", "5000")),
        debug=os.getenv("FLASK_DEBUG", "0") == "1",
    )