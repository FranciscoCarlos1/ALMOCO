import csv
import os
from io import BytesIO, StringIO
from pathlib import Path
from urllib.request import urlopen

from flask import Blueprint, render_template, request, abort, redirect, url_for, Response
from datetime import date, datetime, timedelta
from openpyxl import Workbook, load_workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import Image as RLImage, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from db import get_conn

bp_admin = Blueprint('admin', __name__)
BASE_DIR = Path(__file__).resolve().parent

ADMIN_TOKEN = os.getenv("ALMOCO_ADMIN_TOKEN", "ifc-sbs")
ALLOWED_IMAGE_EXTENSIONS = {".png", ".jpg", ".jpeg", ".webp", ".gif"}

TURMAS = [
    "TIN I A", "TIN I B", "TIN II", "TIN III",
    "TAI I", "TAI II", "TAI III",
    "TST I", "TST II", "TST III", "SERVIDORES"
]


def parse_iso_date(value: str) -> date:
    return datetime.strptime(value, "%Y-%m-%d").date()


def limpar_cpf(valor: str) -> str:
    return "".join(caractere for caractere in valor if caractere.isdigit())


def normalize_header(value: object) -> str:
    texto = " ".join(str(value or "").strip().lower().split())
    replacements = {
        "á": "a",
        "à": "a",
        "â": "a",
        "ã": "a",
        "é": "e",
        "ê": "e",
        "í": "i",
        "ó": "o",
        "ô": "o",
        "õ": "o",
        "ú": "u",
        "ç": "c",
    }
    for original, novo in replacements.items():
        texto = texto.replace(original, novo)
    return texto


def as_clean_text(value: object) -> str:
    if value is None:
        return ""
    return str(value).strip()


def week_start(given_date: date) -> date:
    return given_date - timedelta(days=given_date.weekday())


def month_bounds(given_date: date) -> tuple[date, date]:
    inicio = given_date.replace(day=1)
    if given_date.month == 12:
        proximo = date(given_date.year + 1, 1, 1)
    else:
        proximo = date(given_date.year, given_date.month + 1, 1)
    return inicio, proximo - timedelta(days=1)


def year_bounds(given_date: date) -> tuple[date, date]:
    return date(given_date.year, 1, 1), date(given_date.year, 12, 31)


def period_bounds(given_date: date, periodo: str) -> tuple[date, date, str]:
    if periodo == "dia":
        return given_date, given_date, "Dia"
    if periodo == "mes":
        inicio, fim = month_bounds(given_date)
        return inicio, fim, "Mês"
    if periodo == "ano":
        inicio, fim = year_bounds(given_date)
        return inicio, fim, "Ano"
    segunda = week_start(given_date)
    sexta = segunda + timedelta(days=4)
    return segunda, sexta, "Semana"


def build_quadro_semana(conn, segunda: date, sexta: date) -> tuple[dict[str, int], list[dict[str, int | str]], int]:
    turma_semana_rows = conn.execute(
        """
        SELECT turma,
               data_almoco,
               SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END) AS sim
        FROM respostas
        WHERE data_almoco BETWEEN ? AND ?
        GROUP BY turma, data_almoco
        ORDER BY turma, data_almoco
        """,
        (segunda.isoformat(), sexta.isoformat()),
    ).fetchall()

    semana_sim: dict[str, int] = {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0}
    week_map = {
        segunda.isoformat(): "seg",
        (segunda + timedelta(days=1)).isoformat(): "ter",
        (segunda + timedelta(days=2)).isoformat(): "qua",
        (segunda + timedelta(days=3)).isoformat(): "qui",
        (segunda + timedelta(days=4)).isoformat(): "sex",
    }
    turma_semana = {
        turma: {"seg": 0, "ter": 0, "qua": 0, "qui": 0, "sex": 0, "total": 0}
        for turma in TURMAS
    }

    for row in turma_semana_rows:
        turma = row["turma"]
        dia = week_map.get(str(row["data_almoco"]))
        if turma not in turma_semana or not dia:
            continue
        turma_semana[turma][dia] = row["sim"] or 0

    quadro_importado_rows = conn.execute(
        """
        SELECT turma, data_almoco, sim
        FROM quadro_importado
        WHERE data_almoco BETWEEN ? AND ?
        """,
        (segunda.isoformat(), sexta.isoformat()),
    ).fetchall()

    for row in quadro_importado_rows:
        turma = row["turma"]
        dia = week_map.get(str(row["data_almoco"]))
        if turma not in turma_semana or not dia:
            continue
        turma_semana[turma][dia] = max(turma_semana[turma][dia], max(0, int(row["sim"] or 0)))

    for turma in TURMAS:
        item = turma_semana[turma]
        item["total"] = item["seg"] + item["ter"] + item["qua"] + item["qui"] + item["sex"]
        for dia in semana_sim:
            semana_sim[dia] += item[dia]

    quadro_rows = []
    for idx, turma in enumerate(TURMAS, start=1):
        item = turma_semana[turma]
        quadro_rows.append(
            {
                "ordem": idx,
                "turma_nome": turma,
                "seg": item["seg"],
                "ter": item["ter"],
                "qua": item["qua"],
                "qui": item["qui"],
                "sex": item["sex"],
                "total": item["total"],
            }
        )

    return semana_sim, quadro_rows, sum(semana_sim.values())


def build_respostas_semana(conn, segunda: date, sexta: date) -> list[dict[str, str | int]]:
    respostas_semana_rows = conn.execute(
        """
        SELECT nome, matricula, turma, data_almoco, intencao
        FROM respostas
        WHERE data_almoco BETWEEN ? AND ?
        ORDER BY turma, nome, data_almoco
        """,
        (segunda.isoformat(), sexta.isoformat()),
    ).fetchall()

    respostas_por_pessoa: dict[str, dict[str, str | dict[str, bool]]] = {}
    week_map_respostas = {
        segunda.isoformat(): "seg",
        (segunda + timedelta(days=1)).isoformat(): "ter",
        (segunda + timedelta(days=2)).isoformat(): "qua",
        (segunda + timedelta(days=3)).isoformat(): "qui",
        (segunda + timedelta(days=4)).isoformat(): "sex",
    }
    for row in respostas_semana_rows:
        matricula = row["matricula"]
        if matricula not in respostas_por_pessoa:
            respostas_por_pessoa[matricula] = {
                "nome": row["nome"],
                "turma": row["turma"],
                "dias": {"seg": False, "ter": False, "qua": False, "qui": False, "sex": False},
            }
        dia = week_map_respostas.get(str(row["data_almoco"]))
        if dia and row["intencao"] == "SIM":
            respostas_por_pessoa[matricula]["dias"][dia] = True

    dias_label = {"seg": "Seg", "ter": "Ter", "qua": "Qua", "qui": "Qui", "sex": "Sex"}
    respostas = []
    for item in sorted(respostas_por_pessoa.values(), key=lambda x: (x["turma"], x["nome"])):
        dias = item["dias"]
        checks = [f"{dias_label[dia]} ✅" for dia in ["seg", "ter", "qua", "qui", "sex"] if dias[dia]]
        total_checks = len(checks)
        respostas.append(
            {
                "nome": item["nome"],
                "turma": item["turma"],
                "intencao": " | ".join(checks) if checks else "Sem check na semana",
                "total_checks": total_checks,
            }
        )

    return respostas


def build_respostas_dia(conn, data_filtro: str, intencao_filtro: str) -> list[dict[str, str]]:
    if intencao_filtro in {"SIM", "NAO"}:
        rows = conn.execute(
            """
            SELECT nome, turma, intencao
            FROM respostas
            WHERE data_almoco = ? AND intencao = ?
            ORDER BY turma, nome
            """,
            (data_filtro, intencao_filtro),
        ).fetchall()
    else:
        rows = conn.execute(
            """
            SELECT nome, turma, intencao
            FROM respostas
            WHERE data_almoco = ?
            ORDER BY turma, nome
            """,
            (data_filtro,),
        ).fetchall()

    return [
        {
            "nome": row["nome"],
            "turma": row["turma"],
            "intencao": row["intencao"],
        }
        for row in rows
    ]

@bp_admin.route("/admin")
def admin():
    token = request.args.get("token", "")
    if token != ADMIN_TOKEN:
        abort(403, "Acesso negado. Informe um token válido na URL.")
    data_filtro = request.args.get("data") or date.today().isoformat()
    try:
        data_base = parse_iso_date(data_filtro)
    except ValueError:
        data_base = date.today()
        data_filtro = data_base.isoformat()

    segunda = week_start(data_base)
    sexta = segunda + timedelta(days=4)
    periodo = request.args.get("periodo", "semana").strip().lower()
    if periodo not in {"dia", "semana", "mes", "ano"}:
        periodo = "semana"
    intencao_filtro = request.args.get("intencao", "TODOS").strip().upper()
    if intencao_filtro not in {"TODOS", "SIM", "NAO"}:
        intencao_filtro = "TODOS"
    periodo_inicio, periodo_fim, periodo_label = period_bounds(data_base, periodo)
    cardapio_salvo = request.args.get("cardapio_salvo") == "1"

    with get_conn() as conn:
        cardapio = conn.execute(
            """
            SELECT descricao, imagem_blob
            FROM cardapios
            WHERE data_almoco = ?
            """,
            (data_filtro,),
        ).fetchone()
        resumo_rows = conn.execute(
            """
            SELECT turma,
                   SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END) AS sim,
                   SUM(CASE WHEN intencao = 'NAO' THEN 1 ELSE 0 END) AS nao,
                   SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END) AS total
            FROM respostas
            WHERE data_almoco = ?
            GROUP BY turma
            ORDER BY turma
            """,
            (data_filtro,),
        ).fetchall()
        relatorio_periodo_rows = conn.execute(
            """
            SELECT data_almoco,
                   SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END) AS sim,
                   SUM(CASE WHEN intencao = 'NAO' THEN 1 ELSE 0 END) AS nao
            FROM respostas
            WHERE data_almoco BETWEEN ? AND ?
            GROUP BY data_almoco
            ORDER BY data_almoco
            """,
            (periodo_inicio.isoformat(), periodo_fim.isoformat()),
        ).fetchall()
        total_semana_periodo = conn.execute(
            """
            SELECT COALESCE(SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END), 0) AS total
            FROM respostas
            WHERE data_almoco BETWEEN ? AND ?
            """,
            (segunda.isoformat(), sexta.isoformat()),
        ).fetchone()["total"]
        mes_inicio, mes_fim = month_bounds(data_base)
        total_mes_periodo = conn.execute(
            """
            SELECT COALESCE(SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END), 0) AS total
            FROM respostas
            WHERE data_almoco BETWEEN ? AND ?
            """,
            (mes_inicio.isoformat(), mes_fim.isoformat()),
        ).fetchone()["total"]
        ano_inicio, ano_fim = year_bounds(data_base)
        total_ano_periodo = conn.execute(
            """
            SELECT COALESCE(SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END), 0) AS total
            FROM respostas
            WHERE data_almoco BETWEEN ? AND ?
            """,
            (ano_inicio.isoformat(), ano_fim.isoformat()),
        ).fetchone()["total"]
        respostas_dia = build_respostas_dia(conn, data_filtro, intencao_filtro)

    resumo = {turma: {"sim": 0, "nao": 0, "total": 0} for turma in TURMAS}
    for row in resumo_rows:
        resumo[row["turma"]] = {
            "sim": row["sim"] or 0,
            "nao": row["nao"] or 0,
            "total": row["total"] or 0,
        }

    total_sim = sum(item["sim"] for item in resumo.values())
    total_nao = sum(item["nao"] for item in resumo.values())
    total_geral = total_sim
    total_respostas_dia = total_sim
    total_respostas_filtradas = len(respostas_dia)

    quadro_dia_rows = []
    for idx, turma in enumerate(TURMAS, start=1):
        item = resumo[turma]
        quadro_dia_rows.append(
            {
                "ordem": idx,
                "turma_nome": turma,
                "sim": item["sim"],
                "nao": item["nao"],
                "total": item["sim"],
            }
        )

    total_periodo_sim = 0
    total_periodo_nao = 0
    relatorio_por_dia = []
    for row in relatorio_periodo_rows:
        sim = row["sim"] or 0
        nao = row["nao"] or 0
        total = sim + nao
        total_periodo_sim += sim
        total_periodo_nao += nao
        relatorio_por_dia.append(
            {
                "data": str(row["data_almoco"]),
                "sim": sim,
                "nao": nao,
                "total": total,
            }
        )

    return render_template(
        "admin.html",
        resumo=resumo,
        token=token,
        data_filtro=data_filtro,
        periodo=periodo,
        intencao_filtro=intencao_filtro,
        importado=request.args.get("importado") == "1",
        import_error=request.args.get("import_error"),
        biometria_salva=request.args.get("biometria_salva") == "1",
        biometria_error=request.args.get("biometria_error"),
        importado_quadro=request.args.get("importado_quadro") == "1",
        import_quadro_error=request.args.get("import_quadro_error"),
        backup_restaurado=request.args.get("backup_restaurado") == "1",
        backup_restore_file=request.args.get("backup_restore_file"),
        backup_restore_error=request.args.get("backup_restore_error"),
        backup_manual=request.args.get("backup_manual") == "1",
        backup_manual_error=request.args.get("backup_manual_error"),
        total_sim=total_sim,
        total_nao=total_nao,
        total_geral=total_geral,
        periodo_label=periodo_label,
        periodo_inicio=periodo_inicio.isoformat(),
        periodo_fim=periodo_fim.isoformat(),
        total_semana_periodo=total_semana_periodo,
        total_mes_periodo=total_mes_periodo,
        total_ano_periodo=total_ano_periodo,
        total_periodo_sim=total_periodo_sim,
        total_periodo_nao=total_periodo_nao,
        quadro_dia_rows=quadro_dia_rows,
        respostas_dia=respostas_dia,
        total_respostas_dia=total_respostas_dia,
        total_respostas_filtradas=total_respostas_filtradas,
        relatorio_por_dia=relatorio_por_dia,
        cardapio_texto=cardapio["descricao"] if cardapio else "",
        cardapio_salvo=cardapio_salvo,
        cardapio_imagem=data_filtro if cardapio and cardapio["imagem_blob"] else None,
        cardapio_url=url_for("admin.painel_cardapio", token=token, data=data_filtro),
    )


def _validar_token(token: str) -> None:
    if token != ADMIN_TOKEN:
        abort(403, "Acesso negado. Informe um token válido na URL.")


def _obter_intencao_filtro() -> str:
    intencao_filtro = request.args.get("intencao", "TODOS").strip().upper()
    if intencao_filtro not in {"TODOS", "SIM", "NAO"}:
        return "TODOS"
    return intencao_filtro


def _obter_cardapio(data_filtro: str):
    with get_conn() as conn:
        return conn.execute(
            """
            SELECT descricao, imagem_blob, imagem_mime
            FROM cardapios
            WHERE data_almoco = ?
            """,
            (data_filtro,),
        ).fetchone()


def _buscar_aluno_por_cadastro(conn, matricula: str, cpf: str, identificador_biometrico: str):
    if matricula:
        aluno = conn.execute(
            """
            SELECT matricula, cpf, identificador_biometrico
            FROM alunos
            WHERE matricula = ?
            """,
            (matricula,),
        ).fetchone()
        if aluno:
            return aluno

    if cpf:
        aluno = conn.execute(
            """
            SELECT matricula, cpf, identificador_biometrico
            FROM alunos
            WHERE cpf = ?
            """,
            (cpf,),
        ).fetchone()
        if aluno:
            return aluno

    if identificador_biometrico:
        aluno = conn.execute(
            """
            SELECT matricula, cpf, identificador_biometrico
            FROM alunos
            WHERE identificador_biometrico = ?
            """,
            (identificador_biometrico,),
        ).fetchone()
        if aluno:
            return aluno

    return None


def _salvar_imagem_cardapio(arquivo) -> tuple[bytes, str]:
    nome_arquivo = arquivo.filename or ""
    extensao = os.path.splitext(nome_arquivo)[1].lower()
    if extensao not in ALLOWED_IMAGE_EXTENSIONS:
        raise ValueError("Envie uma imagem PNG, JPG, JPEG, WEBP ou GIF.")

    mime_type = (arquivo.mimetype or "").strip() or {
        ".png": "image/png",
        ".jpg": "image/jpeg",
        ".jpeg": "image/jpeg",
        ".webp": "image/webp",
        ".gif": "image/gif",
    }[extensao]
    return arquivo.read(), mime_type


@bp_admin.post("/admin/importar_alunos")
def importar_alunos():
    token = request.form.get("token", "")
    _validar_token(token)

    arquivo = request.files.get("arquivo_csv")
    data_filtro = request.form.get("data", "")

    if not arquivo or not arquivo.filename:
        return redirect(url_for("admin.admin", token=token, data=data_filtro, import_error="Selecione um arquivo CSV ou XLSX."))

    nome_arquivo = (arquivo.filename or "").lower()
    rows_for_import: list[dict[str, str]] = []

    try:
        if nome_arquivo.endswith(".xlsx"):
            workbook = load_workbook(arquivo.stream, data_only=True)
            sheet = workbook.active
            raw_rows = list(sheet.iter_rows(values_only=True))

            if not raw_rows:
                return redirect(url_for("admin.admin", token=token, data=data_filtro, import_error="Planilha XLSX vazia."))

            header_original = [as_clean_text(col) for col in raw_rows[0]]
            if not any(header_original):
                return redirect(url_for("admin.admin", token=token, data=data_filtro, import_error="XLSX sem cabeçalho."))

            header_normalized = [normalize_header(col) for col in header_original]
            for values in raw_rows[1:]:
                item = {
                    header_normalized[i]: as_clean_text(values[i])
                    for i in range(min(len(header_normalized), len(values)))
                }
                rows_for_import.append(item)
        else:
            payload = arquivo.stream.read().decode("utf-8-sig")
            sample = payload[:2048]
            dialect = csv.Sniffer().sniff(sample, delimiters=",;") if sample.strip() else csv.get_dialect("excel")
            reader = csv.DictReader(StringIO(payload), dialect=dialect)

            if not reader.fieldnames:
                return redirect(url_for("admin.admin", token=token, data=data_filtro, import_error="CSV sem cabeçalho."))

            header_normalized = [normalize_header(item) for item in reader.fieldnames]
            for row in reader:
                item = {
                    header_normalized[i]: as_clean_text(row.get(reader.fieldnames[i]))
                    for i in range(len(header_normalized))
                }
                rows_for_import.append(item)
    except Exception:
        return redirect(url_for("admin.admin", token=token, data=data_filtro, import_error="Não foi possível ler o arquivo (use CSV ou XLSX válido)."))

    if not rows_for_import:
        return redirect(url_for("admin.admin", token=token, data=data_filtro, import_error="Arquivo sem dados para importar."))

    all_keys: set[str] = set()
    for row in rows_for_import:
        all_keys.update(row.keys())

    nome_key = next((h for h in ["nome", "aluno", "nome completo"] if h in all_keys), None)
    matricula_key = next((h for h in ["matricula", "matricula aluno", "ra"] if h in all_keys), None)
    turma_key = next((h for h in ["turma", "serie", "classe"] if h in all_keys), None)
    cpf_key = next((h for h in ["cpf", "cpf aluno"] if h in all_keys), None)
    biometria_key = next((h for h in ["identificador biometrico", "identificador_biometrico", "biometria", "id biometrico", "codigo biometrico"] if h in all_keys), None)

    if not nome_key or not matricula_key or not turma_key:
        return redirect(
            url_for(
                "admin.admin",
                token=token,
                data=data_filtro,
                import_error="Arquivo precisa das colunas nome, matricula e turma. CPF e identificador biométrico são opcionais.",
            )
        )

    importados = 0
    with get_conn() as conn:
        for row in rows_for_import:
            nome = as_clean_text(row.get(nome_key))
            matricula = as_clean_text(row.get(matricula_key))
            turma = as_clean_text(row.get(turma_key))
            cpf = limpar_cpf(as_clean_text(row.get(cpf_key))) if cpf_key else ""
            identificador_biometrico = as_clean_text(row.get(biometria_key)) if biometria_key else ""

            if not nome or not matricula or turma not in TURMAS:
                continue

            conn.execute(
                """
                INSERT INTO alunos (matricula, nome, turma, cpf, identificador_biometrico)
                VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(matricula)
                DO UPDATE SET
                    nome = excluded.nome,
                    turma = excluded.turma,
                    cpf = COALESCE(excluded.cpf, alunos.cpf),
                    identificador_biometrico = COALESCE(excluded.identificador_biometrico, alunos.identificador_biometrico),
                    atualizado_em = CURRENT_TIMESTAMP
                """,
                (matricula, nome, turma, cpf or None, identificador_biometrico or None),
            )
            importados += 1
        conn.commit()

    if importados == 0:
        return redirect(url_for("admin.admin", token=token, data=data_filtro, import_error="Nenhuma linha válida foi importada. Verifique turma, matrícula e nome."))

    return redirect(url_for("admin.admin", token=token, data=data_filtro, importado=1))


@bp_admin.get("/export.csv")
def export_csv() -> Response:
    token = request.args.get("token", "")
    _validar_token(token)

    data_filtro = request.args.get("data") or date.today().isoformat()
    intencao_filtro = _obter_intencao_filtro()
    with get_conn() as conn:
        if intencao_filtro in {"SIM", "NAO"}:
            rows = conn.execute(
                """
                SELECT nome, matricula, turma, data_almoco, intencao, criado_em
                FROM respostas
                WHERE data_almoco = ? AND intencao = ?
                ORDER BY turma, nome
                """,
                (data_filtro, intencao_filtro),
            ).fetchall()
        else:
            rows = conn.execute(
                """
                SELECT nome, matricula, turma, data_almoco, intencao, criado_em
                FROM respostas
                WHERE data_almoco = ?
                ORDER BY turma, nome
                """,
                (data_filtro,),
            ).fetchall()

    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["nome", "matricula", "turma", "data_almoco", "intencao", "criado_em"])
    for row in rows:
        writer.writerow([row["nome"], row["matricula"], row["turma"], row["data_almoco"], row["intencao"], row["criado_em"]])

    csv_data = output.getvalue()
    output.close()
    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=almoco_{data_filtro}_{intencao_filtro.lower()}.csv"},
    )


@bp_admin.get("/export_quadro.csv")
def export_quadro_csv() -> Response:
    token = request.args.get("token", "")
    _validar_token(token)

    data_filtro = request.args.get("data") or date.today().isoformat()
    intencao_filtro = _obter_intencao_filtro()
    try:
        data_base = parse_iso_date(data_filtro)
    except ValueError:
        data_base = date.today()

    with get_conn() as conn:
        resumo_rows = conn.execute(
            """
            SELECT turma,
                   SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END) AS sim
            FROM respostas
            WHERE data_almoco = ?
            GROUP BY turma
            ORDER BY turma
            """,
            (data_filtro,),
        ).fetchall()
        respostas = build_respostas_dia(conn, data_filtro, intencao_filtro)

    resumo = {turma: 0 for turma in TURMAS}
    for row in resumo_rows:
        resumo[row["turma"]] = row["sim"] or 0

    output = StringIO()
    writer = csv.writer(output, delimiter=';')
    writer.writerow(["Data", data_base.isoformat()])
    writer.writerow(["Filtro", intencao_filtro])
    writer.writerow([])
    writer.writerow(["#", "Turma", "SIM", "Total"])
    total_sim = 0
    for idx, turma in enumerate(TURMAS, start=1):
        sim = resumo[turma]
        total_sim += sim
        writer.writerow([idx, turma, sim, sim])
    writer.writerow(["", "Total", total_sim, total_sim])
    writer.writerow([])
    writer.writerow(["Nome", "Turma", "Intencao"])
    for row in respostas:
        writer.writerow([row["nome"], row["turma"], row["intencao"]])
    writer.writerow(["Total exibido", "", len(respostas)])

    csv_data = output.getvalue()
    output.close()
    return Response(
        csv_data,
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename=quadro_diario_{data_base.isoformat()}_{intencao_filtro.lower()}.csv"},
    )


@bp_admin.get("/export_quadro.xlsx")
def export_quadro_xlsx() -> Response:
    token = request.args.get("token", "")
    _validar_token(token)

    data_filtro = request.args.get("data") or date.today().isoformat()
    intencao_filtro = _obter_intencao_filtro()
    try:
        data_base = parse_iso_date(data_filtro)
    except ValueError:
        data_base = date.today()

    with get_conn() as conn:
        resumo_rows = conn.execute(
            """
            SELECT turma,
                   SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END) AS sim
            FROM respostas
            WHERE data_almoco = ?
            GROUP BY turma
            ORDER BY turma
            """,
            (data_filtro,),
        ).fetchall()
        respostas = build_respostas_dia(conn, data_filtro, intencao_filtro)

    resumo = {turma: 0 for turma in TURMAS}
    for row in resumo_rows:
        resumo[row["turma"]] = row["sim"] or 0

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "quadro_diario"
    sheet.append(["#", "Turma", "SIM", "Total"])
    total_sim = 0
    for idx, turma in enumerate(TURMAS, start=1):
        sim = resumo[turma]
        total_sim += sim
        sheet.append([idx, turma, sim, sim])
    sheet.append(["", "Total", total_sim, total_sim])

    respostas_sheet = workbook.create_sheet("respostas_dia")
    respostas_sheet.append(["Nome", "Turma", "Intencao"])
    for row in respostas:
        respostas_sheet.append([row["nome"], row["turma"], row["intencao"]])
    respostas_sheet.append(["Total exibido", "", len(respostas)])

    meta = workbook.create_sheet("metadados")
    meta.append(["campo", "valor"])
    meta.append(["data_referencia", data_base.isoformat()])
    meta.append(["filtro_intencao", intencao_filtro])
    meta.append(["data_referencia", data_filtro])
    meta.append(["gerado_em", datetime.now().isoformat(timespec="seconds")])

    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=quadro_diario_{data_base.isoformat()}_{intencao_filtro.lower()}.xlsx"},
    )


@bp_admin.get("/export_quadro.pdf")
def export_quadro_pdf() -> Response:
    token = request.args.get("token", "")
    _validar_token(token)

    data_filtro = request.args.get("data") or date.today().isoformat()
    intencao_filtro = _obter_intencao_filtro()
    try:
        data_base = parse_iso_date(data_filtro)
    except ValueError:
        data_base = date.today()

    with get_conn() as conn:
        resumo_rows = conn.execute(
            """
            SELECT turma,
                   SUM(CASE WHEN intencao = 'SIM' THEN 1 ELSE 0 END) AS sim,
                   SUM(CASE WHEN intencao = 'NAO' THEN 1 ELSE 0 END) AS nao
            FROM respostas
            WHERE data_almoco = ?
            GROUP BY turma
            ORDER BY turma
            """,
            (data_filtro,),
        ).fetchall()
        respostas = build_respostas_dia(conn, data_filtro, intencao_filtro)

    resumo = {turma: {"sim": 0, "nao": 0} for turma in TURMAS}
    for row in resumo_rows:
        resumo[row["turma"]] = {
            "sim": row["sim"] or 0,
            "nao": row["nao"] or 0,
        }

    quadro_rows = []
    for idx, turma in enumerate(TURMAS, start=1):
        item = resumo[turma]
        quadro_rows.append(
            {
                "ordem": idx,
                "turma_nome": turma,
                "sim": item["sim"],
                "total": item["sim"],
            }
        )

    total_sim = sum(item["sim"] for item in resumo.values())

    pdf_buffer = BytesIO()
    document = SimpleDocTemplate(pdf_buffer, pagesize=landscape(A4), leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    styles = getSampleStyleSheet()
    story = []

    logo_image = None
    logo_url = request.url_root.rstrip("/") + url_for("static", filename="logo_ifc_horizontal_SaoBentodosul.png")
    try:
        with urlopen(logo_url, timeout=5) as response:
            logo_bytes = response.read()
        logo_image = RLImage(BytesIO(logo_bytes), width=320, height=105)
    except Exception:
        logo_path = BASE_DIR / "static" / "logo_ifc_horizontal_SaoBentodosul.png"
        if logo_path.exists():
            try:
                logo_image = RLImage(str(logo_path), width=320, height=105)
            except Exception:
                logo_image = None

    if logo_image is not None:
        logo_image.hAlign = "CENTER"
        story.append(logo_image)
        story.append(Spacer(1, 10))

    story.extend([
        Paragraph("Quadro diario por turma", styles["Title"]),
        Spacer(1, 8),
        Paragraph(f"Data: {data_base.isoformat()}", styles["Normal"]),
        Paragraph(f"Filtro de intencao: {intencao_filtro}", styles["Normal"]),
        Spacer(1, 12),
    ])

    table_data = [["#", "Turma", "SIM", "Total"]]
    for row in quadro_rows:
        table_data.append([row["ordem"], row["turma_nome"], row["sim"], row["total"]])
    table_data.append(["", "Total", total_sim, total_sim])

    table = Table(table_data, colWidths=[40, 420, 120, 120], repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#E8E8E8")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF200")),
        ("ALIGN", (0, 0), (0, -1), "CENTER"),
        ("ALIGN", (2, 0), (-1, -1), "CENTER"),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("GRID", (0, 0), (-1, -1), 0.8, colors.HexColor("#5F5F5F")),
    ]))
    story.append(table)

    story.extend([
        Spacer(1, 18),
        Paragraph("Respostas do dia", styles["Heading2"]),
        Spacer(1, 8),
    ])

    respostas_table_data = [["Nome", "Turma", "Intencao"]]
    for row in respostas:
        respostas_table_data.append([
            row["nome"],
            row["turma"],
            row["intencao"],
        ])
    if len(respostas_table_data) == 1:
        respostas_table_data.append(["Sem respostas no dia", "-", "-"])

    respostas_table_data.append(["Total exibido", "", len(respostas)])

    respostas_table = Table(respostas_table_data, colWidths=[360, 150, 250], repeatRows=1)
    respostas_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#F0F3F7")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTNAME", (0, -1), (-1, -1), "Helvetica-Bold"),
        ("BACKGROUND", (0, -1), (-1, -1), colors.HexColor("#FFF200")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("ALIGN", (-1, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#D0D7DE")),
        ("ROWBACKGROUNDS", (0, 1), (-1, -2), [colors.white, colors.HexColor("#FAFBFC")]),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
        ("LEADING", (0, 0), (-1, -1), 11),
    ]))
    story.append(respostas_table)
    document.build(story)
    pdf_buffer.seek(0)

    return Response(
        pdf_buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=quadro_diario_{data_base.isoformat()}_{intencao_filtro.lower()}.pdf"},
    )


@bp_admin.route("/admin/cardapio", methods=["GET", "POST"])
def painel_cardapio():
    token = request.values.get("token", "")
    _validar_token(token)

    data_filtro = request.values.get("data", "").strip() or date.today().isoformat()

    if request.method == "POST":
        descricao = request.form.get("descricao", "").strip()
        remover_imagem = request.form.get("remover_imagem") == "1"
        arquivo = request.files.get("imagem")
        cardapio_atual = _obter_cardapio(data_filtro)
        imagem_blob = cardapio_atual["imagem_blob"] if cardapio_atual else None
        imagem_mime = cardapio_atual["imagem_mime"] if cardapio_atual else None

        try:
            if arquivo and arquivo.filename:
                imagem_blob, imagem_mime = _salvar_imagem_cardapio(arquivo)
            elif remover_imagem:
                imagem_blob = None
                imagem_mime = None
        except ValueError as exc:
            return render_template(
                "cardapio_admin.html",
                token=token,
                data_filtro=data_filtro,
                cardapio_texto=descricao,
                cardapio_imagem=data_filtro if imagem_blob else None,
                cardapio_salvo=False,
                erro_cardapio=str(exc),
            )

        with get_conn() as conn:
            if descricao or imagem_blob:
                conn.execute(
                    """
                    INSERT INTO cardapios (data_almoco, descricao, imagem_blob, imagem_mime, atualizado_em)
                    VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ON CONFLICT(data_almoco) DO UPDATE SET
                        descricao = excluded.descricao,
                        imagem_blob = excluded.imagem_blob,
                        imagem_mime = excluded.imagem_mime,
                        atualizado_em = CURRENT_TIMESTAMP
                    """,
                    (data_filtro, descricao, imagem_blob, imagem_mime),
                )
            else:
                conn.execute(
                    "DELETE FROM cardapios WHERE data_almoco = ?",
                    (data_filtro,),
                )
            conn.commit()

        return redirect(url_for("admin.painel_cardapio", token=token, data=data_filtro, salvo=1))

    cardapio = _obter_cardapio(data_filtro)
    return render_template(
        "cardapio_admin.html",
        token=token,
        data_filtro=data_filtro,
        cardapio_texto=cardapio["descricao"] if cardapio else "",
        cardapio_imagem=data_filtro if cardapio and cardapio["imagem_blob"] else None,
        cardapio_salvo=request.args.get("salvo") == "1",
        erro_cardapio=None,
    )


@bp_admin.post("/admin/aluno-biometria")
def salvar_aluno_biometria():
    token = request.form.get("token", "")
    _validar_token(token)

    data_filtro = request.form.get("data", "").strip() or date.today().isoformat()
    nome = request.form.get("nome", "").strip()
    matricula = request.form.get("matricula", "").strip()
    turma = request.form.get("turma", "").strip()
    cpf = limpar_cpf(request.form.get("cpf", "").strip())
    identificador_biometrico = request.form.get("identificador_biometrico", "").strip()

    if not nome or not matricula or not turma:
        return redirect(url_for(
            "admin.admin",
            token=token,
            data=data_filtro,
            biometria_error="Informe nome, matrícula e turma.",
        ))

    if turma not in TURMAS:
        return redirect(url_for(
            "admin.admin",
            token=token,
            data=data_filtro,
            biometria_error="Selecione uma turma válida.",
        ))

    if not cpf and not identificador_biometrico:
        return redirect(url_for(
            "admin.admin",
            token=token,
            data=data_filtro,
            biometria_error="Informe CPF ou identificador biométrico para integrar a biometria.",
        ))

    with get_conn() as conn:
        existente = _buscar_aluno_por_cadastro(conn, matricula, cpf, identificador_biometrico)
        if existente and existente["matricula"] != matricula:
            return redirect(url_for(
                "admin.admin",
                token=token,
                data=data_filtro,
                biometria_error="CPF ou identificador biométrico já está vinculado a outro aluno.",
            ))

        conn.execute(
            """
            INSERT INTO alunos (matricula, nome, turma, cpf, identificador_biometrico)
            VALUES (?, ?, ?, ?, ?)
            ON CONFLICT(matricula)
            DO UPDATE SET
                nome = excluded.nome,
                turma = excluded.turma,
                cpf = excluded.cpf,
                identificador_biometrico = excluded.identificador_biometrico,
                atualizado_em = CURRENT_TIMESTAMP
            """,
            (matricula, nome, turma, cpf or None, identificador_biometrico or None),
        )
        conn.commit()

    return redirect(url_for(
        "admin.admin",
        token=token,
        data=data_filtro,
        biometria_salva=1,
    ))
