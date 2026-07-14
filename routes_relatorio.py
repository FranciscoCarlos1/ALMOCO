"""
Rotas para relatórios completos com todos os dados de almoço.
Exporta em CSV, XLSX e PDF com informações detalhadas.
"""
import csv
import os
from datetime import date
from io import BytesIO, StringIO
from pathlib import Path

from flask import Blueprint, request, jsonify, Response
from openpyxl import Workbook
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Spacer, Table, TableStyle, Paragraph
from urllib.request import urlopen

from db import get_conn

bp_relatorio = Blueprint('relatorio', __name__)
BASE_DIR = Path(__file__).resolve().parent
ADMIN_TOKEN = os.getenv("ALMOCO_ADMIN_TOKEN", "ifc-sbs")


def parse_iso_date(value: str) -> date:
    return __import__('datetime').datetime.strptime(value, "%Y-%m-%d").date()


@bp_relatorio.get("/admin/relatorio_diario.csv")
def relatorio_diario_csv():
    """
    Relatório completo do dia em CSV.
    Inclui: nome, matrícula, turma, CPF, presença (SIM/NAO), hora do registro.
    """
    token = request.args.get("token", "")
    if token != ADMIN_TOKEN:
        return jsonify({"ok": False, "erro": "Acesso negado"}), 403
    
    data_filtro = request.args.get("data") or date.today().isoformat()
    
    with get_conn() as conn:
        # Buscar todos os alunos cadastrados
        todos_alunos = conn.execute(
            """
            SELECT matricula, nome, turma, cpf
            FROM alunos
            ORDER BY turma, nome
            """
        ).fetchall()
        
        # Buscar respostas do dia
        respostas_dia = conn.execute(
            """
            SELECT matricula, nome, turma, intencao, criado_em
            FROM respostas
            WHERE data_almoco = ?
            ORDER BY turma, nome
            """,
            (data_filtro,)
        ).fetchall()
    
    # Criar mapa de respostas
    respostas_map = {}
    for resp in respostas_dia:
        respostas_map[resp["matricula"]] = resp
    
    # Gerar CSV
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "MATRÍCULA", "NOME", "TURMA", "CPF", 
        "PRESENÇA", "HORA DO REGISTRO"
    ])
    
    for aluno in todos_alunos:
        matricula = aluno["matricula"]
        resposta = respostas_map.get(matricula)
        
        if resposta:
            writer.writerow([
                matricula,
                resposta["nome"],
                resposta["turma"],
                aluno["cpf"] or "",
                resposta["intencao"],
                resposta["criado_em"] or ""
            ])
        else:
            writer.writerow([
                matricula,
                aluno["nome"],
                aluno["turma"],
                aluno["cpf"] or "",
                "NÃO REGISTRADO",
                ""
            ])
    
    csv_data = output.getvalue()
    output.close()
    
    return Response(
        csv_data,
        mimetype="text/csv; charset=utf-8",
        headers={"Content-Disposition": f"attachment; filename=relatorio_diario_{data_filtro}.csv"}
    )


@bp_relatorio.get("/admin/relatorio_diario.xlsx")
def relatorio_diario_xlsx():
    """
    Relatório completo do dia em XLSX.
    Inclui: nome, matrícula, turma, CPF, presença, hora, biometria cadastrada.
    """
    token = request.args.get("token", "")
    if token != ADMIN_TOKEN:
        return jsonify({"ok": False, "erro": "Acesso negado"}), 403
    
    data_filtro = request.args.get("data") or date.today().isoformat()
    
    with get_conn() as conn:
        # Buscar todos os alunos com biometria
        todos_alunos = conn.execute(
            """
            SELECT matricula, nome, turma, cpf, identificador_biometrico
            FROM alunos
            ORDER BY turma, nome
            """
        ).fetchall()
        
        # Buscar respostas do dia
        respostas_dia = conn.execute(
            """
            SELECT matricula, nome, turma, intencao, criado_em
            FROM respostas
            WHERE data_almoco = ?
            ORDER BY turma, nome
            """,
            (data_filtro,)
        ).fetchall()
    
    # Criar mapa de respostas
    respostas_map = {}
    for resp in respostas_dia:
        respostas_map[resp["matricula"]] = resp
    
    # Contar registros
    total_alunos = len(todos_alunos)
    presentes = sum(1 for a in todos_alunos if respostas_map.get(a["matricula"], {}).get("intencao") == "SIM")
    ausentes = sum(1 for a in todos_alunos if respostas_map.get(a["matricula"], {}).get("intencao") == "NAO")
    nao_registrados = total_alunos - presentes - ausentes
    
    # Criar workbook
    workbook = Workbook()
    
    # Aba 1: Resumo
    resumo_sheet = workbook.active
    resumo_sheet.title = "Resumo"
    resumo_sheet.append(["DATA", data_filtro])
    resumo_sheet.append(["TOTAL DE ALUNOS", total_alunos])
    resumo_sheet.append(["PRESENTES (SIM)", presentes])
    resumo_sheet.append(["AUSENTES (NAO)", ausentes])
    resumo_sheet.append(["NÃO REGISTRADOS", nao_registrados])
    resumo_sheet.append([])
    resumo_sheet.append(["% PRESENTES", f"{(presentes/total_alunos*100) if total_alunos > 0 else 0:.1f}%"])
    
    # Aba 2: Detalhes completos
    detalhes_sheet = workbook.create_sheet("Detalhes")
    detalhes_sheet.append([
        "MATRÍCULA", "NOME", "TURMA", "CPF",
        "PRESENÇA", "HORA DO REGISTRO", "UID RFID"
    ])
    
    for aluno in todos_alunos:
        matricula = aluno["matricula"]
        resposta = respostas_map.get(matricula)
        
        if resposta:
            detalhes_sheet.append([
                matricula,
                resposta["nome"],
                resposta["turma"],
                aluno["cpf"] or "",
                resposta["intencao"],
                resposta["criado_em"] or "",
                aluno["identificador_biometrico"] or ""
            ])
        else:
            detalhes_sheet.append([
                matricula,
                aluno["nome"],
                aluno["turma"],
                aluno["cpf"] or "",
                "NÃO REGISTRADO",
                "",
                aluno["identificador_biometrico"] or ""
            ])
    
    # Aba 3: Presentes
    presentes_sheet = workbook.create_sheet("Presentes")
    presentes_sheet.append(["MATRÍCULA", "NOME", "TURMA", "CPF", "HORA"])
    
    for aluno in todos_alunos:
        resposta = respostas_map.get(aluno["matricula"])
        if resposta and resposta["intencao"] == "SIM":
            presentes_sheet.append([
                aluno["matricula"],
                resposta["nome"],
                resposta["turma"],
                aluno["cpf"] or "",
                resposta["criado_em"] or ""
            ])
    
    # Aba 4: Ausentes
    ausentes_sheet = workbook.create_sheet("Ausentes")
    ausentes_sheet.append(["MATRÍCULA", "NOME", "TURMA", "CPF", "HORA"])
    
    for aluno in todos_alunos:
        resposta = respostas_map.get(aluno["matricula"])
        if resposta and resposta["intencao"] == "NAO":
            ausentes_sheet.append([
                aluno["matricula"],
                resposta["nome"],
                resposta["turma"],
                aluno["cpf"] or "",
                resposta["criado_em"] or ""
            ])
    
    # Aba 5: Não registrados
    nao_registrados_sheet = workbook.create_sheet("Não Registrados")
    nao_registrados_sheet.append(["MATRÍCULA", "NOME", "TURMA", "CPF", "UID RFID"])
    
    for aluno in todos_alunos:
        if aluno["matricula"] not in respostas_map:
            nao_registrados_sheet.append([
                aluno["matricula"],
                aluno["nome"],
                aluno["turma"],
                aluno["cpf"] or "",
                aluno["identificador_biometrico"] or ""
            ])
    
    # Salvar em buffer
    buffer = BytesIO()
    workbook.save(buffer)
    buffer.seek(0)
    
    return Response(
        buffer.getvalue(),
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=relatorio_diario_{data_filtro}.xlsx"}
    )


@bp_relatorio.get("/admin/relatorio_diario.pdf")
def relatorio_diario_pdf():
    """
    Relatório completo do dia em PDF.
    Inclui: resumo, lista de presentes, ausentes e não registrados.
    """
    token = request.args.get("token", "")
    if token != ADMIN_TOKEN:
        return jsonify({"ok": False, "erro": "Acesso negado"}), 403
    
    data_filtro = request.args.get("data") or date.today().isoformat()
    
    with get_conn() as conn:
        # Buscar todos os alunos
        todos_alunos = conn.execute(
            """
            SELECT matricula, nome, turma, cpf
            FROM alunos
            ORDER BY turma, nome
            """
        ).fetchall()
        
        # Buscar respostas do dia
        respostas_dia = conn.execute(
            """
            SELECT matricula, nome, turma, intencao, criado_em
            FROM respostas
            WHERE data_almoco = ?
            ORDER BY turma, nome
            """,
            (data_filtro,)
        ).fetchall()
    
    # Criar mapa de respostas
    respostas_map = {}
    for resp in respostas_dia:
        respostas_map[resp["matricula"]] = resp
    
    # Contar
    total_alunos = len(todos_alunos)
    presentes_list = []
    ausentes_list = []
    nao_registrados_list = []
    
    for aluno in todos_alunos:
        resposta = respostas_map.get(aluno["matricula"])
        if resposta:
            if resposta["intencao"] == "SIM":
                presentes_list.append(resposta)
            else:
                ausentes_list.append(resposta)
        else:
            nao_registrados_list.append(aluno)
    
    # Criar PDF
    pdf_buffer = BytesIO()
    doc = SimpleDocTemplate(
        pdf_buffer,
        pagesize=landscape(A4),
        leftMargin=24, rightMargin=24,
        topMargin=24, bottomMargin=24
    )
    
    styles = getSampleStyleSheet()
    story = []
    
    # Título
    story.append(Paragraph("RELATÓRIO DIÁRIO DE ALMOÇO", styles["Title"]))
    story.append(Paragraph(f"Data: {data_filtro}", styles["Normal"]))
    story.append(Spacer(1, 12))
    
    # Resumo
    story.append(Paragraph("RESUMO", styles["Heading2"]))
    resumo_data = [
        ["Total de Alunos", str(total_alunos)],
        ["Presentes (SIM)", str(len(presentes_list))],
        ["Ausentes (NAO)", str(len(ausentes_list))],
        ["Não Registrados", str(len(nao_registrados_list))],
        ["% Presentes", f"{(len(presentes_list)/total_alunos*100) if total_alunos > 0 else 0:.1f}%"]
    ]
    resumo_table = Table(resumo_data, colWidths=[400, 200])
    resumo_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0066CC")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, 0), 12),
        ("ALIGN", (0, 0), (-1, -1), "CENTER"),
        ("GRID", (0, 0), (-1, -1), 1, colors.black),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0F0F0")])
    ]))
    story.append(resumo_table)
    story.append(Spacer(1, 12))
    
    # Presentes
    if presentes_list:
        story.append(Paragraph(f"PRESENTES ({len(presentes_list)})", styles["Heading2"]))
        presentes_data = [["MATRÍCULA", "NOME", "TURMA", "HORA"]]
        for resp in presentes_list:
            presentes_data.append([
                resp["matricula"],
                resp["nome"],
                resp["turma"],
                resp["criado_em"] or ""
            ])
        presentes_table = Table(presentes_data, colWidths=[150, 350, 150, 200])
        presentes_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#00AA00")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F0FFF0")])
        ]))
        story.append(presentes_table)
        story.append(Spacer(1, 12))
    
    # Ausentes
    if ausentes_list:
        story.append(Paragraph(f"AUSENTES ({len(ausentes_list)})", styles["Heading2"]))
        ausentes_data = [["MATRÍCULA", "NOME", "TURMA", "HORA"]]
        for resp in ausentes_list:
            ausentes_data.append([
                resp["matricula"],
                resp["nome"],
                resp["turma"],
                resp["criado_em"] or ""
            ])
        ausentes_table = Table(ausentes_data, colWidths=[150, 350, 150, 200])
        ausentes_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#CC0000")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFF0F0")])
        ]))
        story.append(ausentes_table)
        story.append(Spacer(1, 12))
    
    # Não registrados
    if nao_registrados_list:
        story.append(Paragraph(f"NÃO REGISTRADOS ({len(nao_registrados_list)})", styles["Heading2"]))
        nao_reg_data = [["MATRÍCULA", "NOME", "TURMA"]]
        for aluno in nao_registrados_list:
            nao_reg_data.append([
                aluno["matricula"],
                aluno["nome"],
                aluno["turma"]
            ])
        nao_reg_table = Table(nao_reg_data, colWidths=[150, 350, 150])
        nao_reg_table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#FFAA00")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.black),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "CENTER"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#FFFFF0")])
        ]))
        story.append(nao_reg_table)
    
    # Gerar PDF
    doc.build(story)
    pdf_buffer.seek(0)
    
    return Response(
        pdf_buffer.getvalue(),
        mimetype="application/pdf",
        headers={"Content-Disposition": f"attachment; filename=relatorio_diario_{data_filtro}.pdf"}
    )


@bp_relatorio.get("/admin/relatorio_json")
def relatorio_json():
    """
    Retorna relatório completo em JSON.
    Útil para integração com sistemas externos.
    """
    token = request.args.get("token", "")
    if token != ADMIN_TOKEN:
        return jsonify({"ok": False, "erro": "Acesso negado"}), 403
    
    data_filtro = request.args.get("data") or date.today().isoformat()
    
    with get_conn() as conn:
        # Buscar todos os alunos
        todos_alunos = conn.execute(
            """
            SELECT matricula, nome, turma, cpf, identificador_biometrico
            FROM alunos
            ORDER BY turma, nome
            """
        ).fetchall()
        
        # Buscar respostas do dia
        respostas_dia = conn.execute(
            """
            SELECT matricula, nome, turma, intencao, criado_em
            FROM respostas
            WHERE data_almoco = ?
            ORDER BY turma, nome
            """,
            (data_filtro,)
        ).fetchall()
    
    # Criar mapa
    respostas_map = {}
    for resp in respostas_dia:
        respostas_map[resp["matricula"]] = resp
    
    # Preparar dados
    presentes = []
    ausentes = []
    nao_registrados = []
    
    for aluno in todos_alunos:
        resposta = respostas_map.get(aluno["matricula"])
        aluno_data = {
            "matricula": aluno["matricula"],
            "nome": aluno["nome"],
            "turma": aluno["turma"],
            "cpf": aluno["cpf"] or "",
            "uid_rfid": aluno["identificador_biometrico"] or ""
        }
        
        if resposta:
            aluno_data["presenca"] = resposta["intencao"]
            aluno_data["hora_registro"] = resposta["criado_em"] or ""
            
            if resposta["intencao"] == "SIM":
                presentes.append(aluno_data)
            else:
                ausentes.append(aluno_data)
        else:
            aluno_data["presenca"] = "NÃO_REGISTRADO"
            aluno_data["hora_registro"] = ""
            nao_registrados.append(aluno_data)
    
    return jsonify({
        "ok": True,
        "data": data_filtro,
        "resumo": {
            "total_alunos": len(todos_alunos),
            "presentes": len(presentes),
            "ausentes": len(ausentes),
            "nao_registrados": len(nao_registrados),
            "percentual_presenca": f"{(len(presentes)/len(todos_alunos)*100) if todos_alunos else 0:.1f}%"
        },
        "presentes": presentes,
        "ausentes": ausentes,
        "nao_registrados": nao_registrados
    })
